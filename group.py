#Python script using openpyxl 
#sorts rows by groups of 30 by category


from openpyxl import *
print("Import Successful")

wb = load_workbook(filename = "8.12 Postcard_Distribution List NEWEST.xlsx")
ws = wb['Group']

bdms = []
num_rows= 0

#Get number of rows and list of BDMs
for row in ws.iter_rows(min_row = 2,min_col = 16,max_col =16,max_row = 26360):
	for cell in row:	
		if cell.value not in bdms and type(cell.value) == str:
			bdms.append(cell.value)
	num_rows += 1	

print(bdms)
#for each person
for bdm in bdms:
	print("Working on:",bdm)
	#reset counter
	temp_counter = 0  #this keeps track of how many rows in 
	group_counter = 1
	for r in range(1,num_rows):
		temp_val = ws.cell(row = r, column = 16).value #column with person name
		if temp_val == bdm:
			if temp_val == ws.cell(row = r-1, column = 16).value: #checking if previous row has same name
				if ws.cell(row = r, column = 18).value != ws.cell(row = r-1, column = 18).value: #category match
					temp_counter = 0
					group_counter = 1
					ws.cell(row = r, column = 14, value = group_counter)
					print(ws.cell(row = r, column = 14))
					temp_counter += 1
				else:
					ws.cell(row = r, column = 14, value = group_counter)
					print(ws.cell(row = r, column = 14))
					temp_counter += 1
					if temp_counter > 29:
						temp_counter = 0
						group_counter += 1
					
			else:
				temp_counter = 0
				group_counter = 1
				ws.cell(row = r, column = 14, value = group_counter)
				temp_counter += 1


wb.save("8.12 Postcard_Distribution List NEWEST.xlsx")
print("saved")
