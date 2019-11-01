import openpyxl
print("import successful")

path = "12.28.pythontest.xlsx" #path to the file to be read

column_names = [ #names of columns to search for
    "First Name",
    "Last Name",
    "Address",
    "City",
    "State",
    "Zip (5)",
    "Salesperson",
    "Vertical",
    "Company",
    "Email Address",
    "Phone Number"
]
column_values = {}

print("make sure there are no blank rows or columns before the table.")

print("loading workbook...")
wb = openpyxl.load_workbook(path)
print("workbook loaded.")
sheet_name = input("data sheet name?")
sheet = wb[sheet_name]
# print(len(sheet[1]))
for col in range(1,len(sheet[1])+1):  #get column_nums for columns
    val = sheet.cell(row=1,column=col).value
    if val in column_names: #if it's a column we're looking for
        temp_array= [] 
        for i in range(0,sheet.max_row-1):
            temp_array.append(sheet.cell(row=i+2,column=col).value)  #stick all it's values into an array
        column_values[val] = temp_array   #add array to dictionary under value name

#creating import wb
new_wb = openpyxl.Workbook()

#creating lead sheet
lead_columns = [
    "Lead #",
    "Lead Name",
    "Third Party Location Code",
    "Third Party Location ID",
    "Notes",
    "Lead Type",
    "Industry",
    "Status",
    "Assigned To(Team Name/Employee#)",
    "Assigned To Type",
    "Address 1",
    "Address 2",
    "City",
    "state",
    "Zipcode",
    "County",
    "School District"
]

contact_columns = [
    "Lead #",
    "Company Name",
    "First Name",
    "Middle Name",
    "Last Name",
    "Primary Email",
    "Secondary Email",
    "Primary Phone",
    "Secondary Phone",
    "Primary Cell #",
    "Secondary Cell #",
    "Fax"
]

branch_columns = ["Lead #","Branch"]
source_columns = ["Lead #","Source"]

crm_codes = {
    "Augustine Beltran":6,
    "Craig Brown":12,
    "Duane Schierbaum":7,
    "Howard Jacobs":8,
    "Richard Ortega":30,
    "Robin Ramirez":26,
    "Sean Campbell":10,
    "Dee Montano":2,
    "Erik Martinez":5,
    "Jason Dejournett":4
}


lead_sheet = new_wb.active
lead_sheet.title = "Lead"

#creating sheets
contacts_sheet = new_wb.create_sheet("Lead Contacts")
branches_sheet = new_wb.create_sheet("Lead Branchs")
source_sheet   = new_wb.create_sheet("Source")



# for idx,val in enumerate(lead_columns): #adding titles
#     lead_sheet.cell(row=1,column=idx+1).value = val

def set_titles(sheet,column_titles):
    for idx,val in enumerate(column_titles): #adding titles
        sheet.cell(row=1,column=idx+1).value = val

set_titles(lead_sheet,lead_columns)
set_titles(contacts_sheet,contact_columns)
set_titles(branches_sheet,branch_columns)
set_titles(source_sheet,source_columns)



###
### FOR LEAD PAGE
###

#for lead name
for i in range(0,len(column_values["Company"])):
    lead_sheet.cell(row=i+2,column=2).value = column_values["Company"][i]

#for lead type
for i in range(0,len(column_values["Company"])):
    lead_sheet.cell(row=i+2,column=6).value = "Commercial"

#for status
for i in range(0,len(column_values["Company"])):
    lead_sheet.cell(row=i+2,column=8).value = "New"

#for assigned to type
for i in range(0,len(column_values["Company"])):
    lead_sheet.cell(row=i+2,column=10).value = "Individual"

#for industry
for i in range(0,len(column_values["Vertical"])):
    lead_sheet.cell(row=i+2,column=7).value = column_values["Vertical"][i]

#for address
for i in range(0,len(column_values["Address"])):
    lead_sheet.cell(row=i+2,column=11).value = column_values["Address"][i]

#for city
for i in range(0,len(column_values["City"])):
    lead_sheet.cell(row=i+2,column=13).value = column_values["City"][i]

#for state
for i in range(0,len(column_values["State"])):
    lead_sheet.cell(row=i+2,column=14).value = column_values["State"][i]

#for zip
for i in range(0,len(column_values["Zip (5)"])):
    lead_sheet.cell(row=i+2,column=15).value = column_values["Zip (5)"][i]

#for notes
for i in range(0,len(column_values["Salesperson"])):
    lead_sheet.cell(row=i+2,column=5).value = crm_codes[column_values["Salesperson"][i]]

###
### FOR CONTACT PAGE
###

#for company name
for i in range(0,len(column_values["Salesperson"])):
    contacts_sheet.cell(row=i+2,column=2).value = column_values["Company"][i]

#for first name
for i in range(0,len(column_values["Salesperson"])):
    contacts_sheet.cell(row=i+2,column=3).value = column_values["First Name"][i]

#for last name
for i in range(0,len(column_values["Salesperson"])):
    contacts_sheet.cell(row=i+2,column=5).value = column_values["Last Name"][i]

#for primary email
for i in range(0,len(column_values["Salesperson"])):
    contacts_sheet.cell(row=i+2,column=6).value = column_values["Email Address"][i]

#for primary phone
for i in range(0,len(column_values["Salesperson"])):
    contacts_sheet.cell(row=i+2,column=8).value = column_values["Phone Number"][i]

###
### FOR BRANCH PAGE
###

#for branch
for i in range(0,len(column_values["Salesperson"])):
    branches_sheet.cell(row=i+2,column=2).value = "Corporate"

###
### FOR SOURCE PAGE
###

#for branch
for i in range(0,len(column_values["Salesperson"])):
    source_sheet.cell(row=i+2,column=2).value = "2019 Postcard Campaign"


#saving workbook
print("saving...")
new_wb.save("test_import_"+sheet_name+".xlsx")
print("saved.")